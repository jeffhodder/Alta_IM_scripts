import imaplib, email
import os.path
from datetime import date, timedelta
import datetime
from dateutil.relativedelta import relativedelta
import os
import glob
from openpyxl import load_workbook
import pyodbc

user = 'j.adjaho@altatrading.com'
password = 'table40!longerchair'
imap_url = 'imap-mail.outlook.com'
attachment_dir = '\\\\altfps\\arcadiagroup$\Midoffice\PVM Dubai'


def next_weekday(d, weekday):
    days_ahead = weekday - d.weekday()
    if days_ahead <= 0:  # Target day already happened this week
        days_ahead += 7
    return d + datetime.timedelta(days_ahead)


# d1 = datetime.date(2021,12, 08)
# next_monday = next_weekday(d1, 0) # 0 = Monday, 1=Tuesday, 2=Wednesday...


todays_date_raw = date.today()
string_date = str(todays_date_raw)
file_year = int(str('20' + string_date[2:4]))
file_month = int(string_date[5:7])
file_day = int(string_date[8:10])
next_monday = next_weekday(datetime.date(file_year, file_month, file_day), 0)
today_date = str(string_date[8:10] + string_date[6:7] + string_date[2:4])
current_date = datetime.date(file_year, file_month, file_day)

conn = pyodbc.connect('DRIVER={SQL Server Native Client 11.0};SERVER=ArcSQL;DATABASE=Traction;TRUSTED_CONNECTION=yes')
cursor = conn.cursor()
cursor.execute("SELECT * FROM view_Holidays")

holidaydate = []
for row in cursor.fetchall():
    if row[2] == 'Platts' or row[2] == 'ICE':
        holidaydate.append(row[0])

if current_date in holidaydate:
    print('Today is a holiday')
    quit()


def get_attachments(msg):
    for part in msg.walk():
        if part.get_content_maintype == 'multipart':
            continue
        if part.get('Content-Disposition') is None:
            continue
        fileName = part.get_filename()

        if bool(fileName):
            filePath = os.path.join(attachment_dir, fileName)
            with open(filePath, 'wb') as f:
                f.write(part.get_payload(decode=True))


con = imaplib.IMAP4_SSL(imap_url)
con.login(user, password)
# print(con.list())


# selecting the first email
email_id_raw = str(con.select('AutomationFolder/PVM_Dubai'))
email_id = email_id_raw[10:-3]
email_id_bytes = bytes(email_id, encoding='utf8')

print(email_id_bytes)
print(con.select('AutomationFolder/PVM_Dubai'))

# this should run after 6pm and select the most recent email, i.e email_id_bytes
result, data = con.fetch(email_id_bytes, '(RFC822)')
raw = email.message_from_bytes(data[0][1])
get_attachments(raw)

newest_file = max(glob.iglob('\\\\altfps\\arcadiagroup$\Midoffice\PVM Dubai/*.xlsx'), key=os.path.getctime)

# file downloaded from email
pvm_dubai_url = max(glob.iglob('\\\\altfps\\arcadiagroup$\Midoffice\PVM Dubai/*.xlsx'), key=os.path.getctime)
pvm_dubai_original = load_workbook(pvm_dubai_url, data_only=True)
pvm_dubai_sheet = pvm_dubai_original['Sheet1']

upload_list = []


def month_function(dubai_month):
    if dubai_month == 1:
        return 'Jan'
    if dubai_month == 2:
        return 'Feb'
    if dubai_month == 3:
        return 'Mar'
    if dubai_month == 4:
        return 'Apr'
    if dubai_month == 5:
        return 'May'
    if dubai_month == 6:
        return 'Jun'
    if dubai_month == 7:
        return 'Jul'
    if dubai_month == 8:
        return 'Aug'
    if dubai_month == 9:
        return 'Sep'
    if dubai_month == 10:
        return 'Oct'
    if dubai_month == 11:
        return 'Nov'
    if dubai_month == 12:
        return 'Dec'


# DBIEFP
if month_function(file_month + 2) in pvm_dubai_sheet['A12'].value:
    current_month = file_month + 2
else:
    current_month = file_month + 3

for i in range(0, 3):
    start_month = datetime.date(file_year, current_month, 1)
    spread_month = start_month + relativedelta(months=i)
    dbiefp = (current_date, 'DBIEFP', spread_month, pvm_dubai_sheet.cell(row=12 + i, column=2).value)
    upload_list.append(dbiefp)

# brent/dubai swaps
if month_function(file_month) in pvm_dubai_sheet['A18'].value:
    current_month = file_month
else:
    current_month = file_month + 1

for i in range(0, 15):
    start_month = datetime.date(file_year, current_month, 1)
    spread_month = start_month + relativedelta(months=i)
    brtdbi = (current_date, 'BOD', spread_month, pvm_dubai_sheet.cell(row=18 + i, column=2).value)
    upload_list.append(brtdbi)

# dubai spreads
if month_function(file_month) in pvm_dubai_sheet['A59'].value:
    current_month = file_month
else:
    current_month = file_month + 1

for i in range(0, 14):
    start_month = datetime.date(file_year, current_month, 1)
    spread_month = start_month + relativedelta(months=i)
    dbispr = (current_date, 'DBISPR', spread_month, pvm_dubai_sheet.cell(row=59 + i, column=2).value)
    upload_list.append(dbispr)

print(upload_list)

conn = pyodbc.connect(
    'DRIVER={SQL Server Native Client 11.0};SERVER=ArcSQL;DATABASE=RiskSandbox;TRUSTED_CONNECTION=yes')
cursor = conn.cursor()
cursor.executemany("INSERT INTO PVMPrices(AsOf, Instrument, ContractDate, Value) VALUES(?,?,?,?)", upload_list)
cursor.commit()

cursor.close()
conn.close()

print('done')