import zipfile
import csv
from openpyxl import load_workbook
from datetime import datetime, timedelta
import pyodbc


yesterday = datetime.today() - timedelta(days=1)
raw_date = yesterday.strftime('%m%d%y')

unwanted_days = [5, 6]
wanted_days = [0, 1, 2, 3, 4]
if yesterday.weekday() in unwanted_days:
    if datetime.today().weekday() == 0:
        last_friday = datetime.today() - timedelta(days=3)
        raw_date_friday = last_friday.strftime('%m%d%y')
        tone_date = str(raw_date_friday[2:4] + raw_date_friday[0:2])
    else:
        quit()
else:
    tone_date = str(raw_date[2:4] + raw_date[0:2])

# uncomment line below for retrospective date
# tone_date = '0703'

file_location = str('\\\\altfps\\arcadiagroup$\Midoffice\Tala IM\\' + tone_date + 'F.zip')

file_location = str('\\\\altfps\\arcadiagroup$\Midoffice\Tala IM\\' + tone_date + 'F.zip')
new_file =str(file_location + '\\' +'MO' + tone_date + 'F')
open_file = str('\\\\altfps\\arcadiagroup$\Midoffice\Tala IM\IM File\\' + 'MO' + tone_date + 'F.csv')

with zipfile.ZipFile(file_location, 'r') as my_zip:
     my_zip.extractall('\\\\altfps\\arcadiagroup$\Midoffice\Tala IM\IM File')

im_automationfile = load_workbook('\\\\altfps\\arcadiagroup$\Midoffice\Tala IM\Tala IM Automation.xlsx')
im_raw_data_tab = im_automationfile['IM_raw_data']

for row in im_raw_data_tab['A:G']:
  for cell in row:
    cell.value = None

initial_margin_amount = []
currency = []
client_code = []
im_date = []
group_code = []
exposure_amount = []
long_ledger_balance = []
with open(open_file,'r') as csv_file:
    csv_reader = csv.reader(csv_file)

    for line in csv_reader:
        group_code.append(line[1])
        client_code.append(line[2])
        currency.append(line[3])
        initial_margin_amount.append(line[7])
        exposure_amount.append(line[9])
        im_date.append(line[10])
        long_ledger_balance.append(line[14])


for i in range(1, len(initial_margin_amount)+1):
    im_raw_data_tab.cell(row=i, column=1).value = client_code[i - 1]
    im_raw_data_tab.cell(row=i, column=2).value = group_code[i - 1]
    im_raw_data_tab.cell(row=i, column=2).value = im_raw_data_tab.cell(row=i, column=2).value.strip()
    im_raw_data_tab.cell(row=i, column=3).value = im_date[i-1]
    im_raw_data_tab.cell(row=i, column=3).value = im_raw_data_tab.cell(row=i, column=3).value.strip()
    im_raw_data_tab.cell(row=i, column=4).value = currency[i-1]
    im_raw_data_tab.cell(row=i, column=4).value = im_raw_data_tab.cell(row=i, column=4).value.strip()
    im_raw_data_tab.cell(row=i, column=5).value = initial_margin_amount[i-1]
    im_raw_data_tab.cell(row=i, column=5).value = im_raw_data_tab.cell(row=i, column=5).value.strip()
    im_raw_data_tab.cell(row=i, column=6).value = exposure_amount[i - 1]
    im_raw_data_tab.cell(row=i, column=6).value = im_raw_data_tab.cell(row=i, column=6).value.strip()
    im_raw_data_tab.cell(row=i, column=7).value = long_ledger_balance[i - 1]
    im_raw_data_tab.cell(row=i, column=7).value = im_raw_data_tab.cell(row=i, column=7).value.strip()


im_automationfile.save('\\\\altfps\\\\arcadiagroup$\Midoffice\Tala IM\Tala IM Automation.xlsx')

conn = pyodbc.connect('DRIVER={SQL Server Native Client 11.0};SERVER=ArcSQL;DATABASE=RiskSandbox;TRUSTED_CONNECTION=yes')
cursor = conn.cursor()

im_automationfile = load_workbook('\\\\altfps\\\\arcadiagroup$\Midoffice\Tala IM\Tala IM Automation.xlsx')
im_raw_data_tab = im_automationfile['IM_raw_data']
asofday = int(im_raw_data_tab.cell(row=2,column=3).value[:2])
asofmonth = int(im_raw_data_tab.cell(row=2,column=3).value[3:5])
asofyear = int(str('20' + im_raw_data_tab.cell(row=2,column=3).value[6:8]))
asofdate_time = datetime(asofyear, asofmonth, asofday)
no_rows= len(initial_margin_amount)

# IM UPLOAD
for i in range(2, no_rows + 1):
        asofdate_time
        if i in range(no_rows - 2,no_rows + 1):
            clearing_account = 'Group'
            value = im_raw_data_tab.cell(row=i, column=5).value
            group_code = im_raw_data_tab.cell(row=i, column=2).value
            currency = im_raw_data_tab.cell(row=i, column=4).value
            if type(im_raw_data_tab.cell(row=i, column=5).value) is not str:
                value = 0
        else:
            group_code = im_raw_data_tab.cell(row=i, column=2).value
            clearing_account = im_raw_data_tab.cell(row=i, column=1).value
            value = im_raw_data_tab.cell(row=i, column=5).value
            currency = im_raw_data_tab.cell(row=i, column=4).value
            if type(im_raw_data_tab.cell(row=i, column=5).value) is not str:
                value = 0

        data = [asofdate_time, clearing_account, value, group_code, currency]
        print(data)


        cursor.execute("INSERT INTO InitialMarginInterest(AsOf, ClearingAccount, Value, GroupCode, Currency) VALUES(?,?,?,?,?)", data)
        cursor.commit()

# CASH BALANCE upload
for k in range(no_rows - 2, no_rows + 1):
    clearing_account_name = str(im_raw_data_tab.cell(row=k, column=4).value + '_exposure')
    data = [asofdate_time, clearing_account_name, im_raw_data_tab.cell(row=k, column=6).value,'TALA', im_raw_data_tab.cell(row=k, column=4).value]
    cursor.execute("INSERT INTO InitialMarginInterest(AsOf, ClearingAccount, Value, GroupCode, Currency) VALUES(?,?,?,?,?)", data)
    cursor.commit()
    print(data)

for k in range(no_rows - 2, no_rows + 1):
    clearing_account_name = str(im_raw_data_tab.cell(row=k, column=4).value + '_cashbal')
    data = [asofdate_time, clearing_account_name, im_raw_data_tab.cell(row=k, column=7).value,'TALA', im_raw_data_tab.cell(row=k, column=4).value]
    cursor.execute("INSERT INTO InitialMarginInterest(AsOf, ClearingAccount, Value, GroupCode, Currency) VALUES(?,?,?,?,?)", data)
    cursor.commit()
    print(data)


cursor.close()
conn.close()
print('done')