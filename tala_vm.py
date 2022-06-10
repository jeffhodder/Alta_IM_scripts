import zipfile
import csv
from openpyxl import load_workbook
from datetime import datetime, timedelta



yesterday = datetime.today() - timedelta(days=1)
raw_date = yesterday.strftime('%m%d%y')

unwanted_days = [5, 6]
wanted_days = [0, 1, 2, 3, 4]
if yesterday.weekday() in unwanted_days:
    if datetime.today().weekday() == 0:
        last_friday = datetime.today() - timedelta(days=3)
        raw_date_friday = last_friday.strftime('%m%d%y')
        tone_date = str(raw_date_friday[2:4] + raw_date_friday[0:2])
    else:
        quit()
else:
    tone_date = str(raw_date[2:4] + raw_date[0:2])

# uncomment line below for retrospective date
tone_date = '0703'

file_location = str('\\\\altfps\\arcadiagroup$\Midoffice\Tala IM\\' + tone_date + 'F.zip')

file_location = str('\\\\altfps\\arcadiagroup$\Midoffice\Tala IM\\' + tone_date + 'F.zip')
new_file =str(file_location + '\\' +'MO' + tone_date + 'F')
open_file = str('\\\\altfps\\arcadiagroup$\Midoffice\Tala IM\IM File\\' + 'MO' + tone_date + 'F.csv')

with zipfile.ZipFile(file_location, 'r') as my_zip:
     my_zip.extractall('\\\\altfps\\arcadiagroup$\Midoffice\Tala IM\IM File')

im_automationfile = load_workbook('\\\\altfps\\\\arcadiagroup$\Midoffice\Tala IM\\variation_margin.xlsx')
im_raw_data_tab = im_automationfile['im_raw_data']


initial_margin_amount = []
variation_margin_amount = []
currency = []
client_code = []
im_date = []
group_code = []
with open(open_file,'r') as csv_file:
    csv_reader = csv.reader(csv_file)

    for line in csv_reader:
        group_code.append(line[1])
        client_code.append(line[2])
        currency.append(line[3])
        initial_margin_amount.append(line[7])
        variation_margin_amount.append(line[6])
        im_date.append(line[10])

for j in range(1,2000):
    if im_raw_data_tab.cell(row=j, column=2).value == None:
        k=j
        break
l=1
for i in range(k, k+len(initial_margin_amount)):
    if l<len(initial_margin_amount):
        im_raw_data_tab.cell(row=i, column=1).value = client_code[l]
        im_raw_data_tab.cell(row=i, column=2).value = group_code[l]
        im_raw_data_tab.cell(row=i, column=2).value = im_raw_data_tab.cell(row=i, column=2).value.strip()
        im_raw_data_tab.cell(row=i, column=3).value = im_date[l]
        im_raw_data_tab.cell(row=i, column=3).value = im_raw_data_tab.cell(row=i, column=3).value.strip()
        im_raw_data_tab.cell(row=i, column=4).value = currency[l]
        im_raw_data_tab.cell(row=i, column=4).value = im_raw_data_tab.cell(row=i, column=4).value.strip()
        im_raw_data_tab.cell(row=i, column=5).value = initial_margin_amount[l]
        im_raw_data_tab.cell(row=i, column=5).value = im_raw_data_tab.cell(row=i, column=5).value.strip()
        im_raw_data_tab.cell(row=i, column=6).value = variation_margin_amount[l]
        im_raw_data_tab.cell(row=i, column=6).value = im_raw_data_tab.cell(row=i, column=6).value.strip()
        l+=1



im_automationfile.save('\\\\altfps\\\\arcadiagroup$\Midoffice\Tala IM\\variation_margin.xlsx')

