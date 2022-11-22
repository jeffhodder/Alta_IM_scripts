import zipfile
import csv
from openpyxl import load_workbook
from datetime import datetime, timedelta,date
import os
from os.path import exists
import pyodbc


yesterday = datetime.today() - timedelta(days=1)
raw_date = yesterday.strftime('%m%d%y')

unwanted_days = [5, 6]
wanted_days = [0, 1, 2, 3, 4]
if yesterday.weekday() in unwanted_days:
    if datetime.today().weekday() == 0:
        last_friday = datetime.today() - timedelta(days=3)
        raw_date_friday = last_friday.strftime('%m%d%y')
        today_date = str(raw_date_friday[2:4] + raw_date_friday[0:2])
    else:
        quit()
else:
    today_date = str(raw_date[2:4] + raw_date[0:2])

# uncomment line below for retrospective date
# today_date = '1210'

file_location = str('\\\\altfps\\arcadiagroup$\Midoffice\IM Attachments\\' + today_date + 'F.zip')

file_location = str('\\\\altfps\\arcadiagroup$\Midoffice\IM Attachments\\' + today_date + 'F.zip')
new_file =str(file_location + '\\' +'MO' + today_date + 'F')
open_file = str('\\\\altfps\\arcadiagroup$\Midoffice\IM Attachments\IM File\\' + 'MO' + today_date + 'F.csv')

with zipfile.ZipFile(file_location, 'r') as my_zip:
     my_zip.extractall('\\\\altfps\\arcadiagroup$\Midoffice\IM Attachments\IM File')

im_automationfile = load_workbook('\\\\altfps\\arcadiagroup$\Midoffice\IM Attachments\IM_Automationfile.xlsm', keep_vba=True)
im_raw_data_tab = im_automationfile['IM_raw_data']

initial_margin_amount = []
client_code = []
im_date = []
group_code = []
with open(open_file,'r') as csv_file:
    csv_reader = csv.reader(csv_file)

    for line in csv_reader:
        initial_margin_amount.append(line[7])
        client_code.append(line[2])
        im_date.append(line[10])
        group_code.append(line[1])

for i in range(1, len(initial_margin_amount)+1):
    im_raw_data_tab.cell(row=i, column=5).value = initial_margin_amount[i-1]
    im_raw_data_tab.cell(row=i, column=5).value = im_raw_data_tab.cell(row=i, column=5).value.strip()
    im_raw_data_tab.cell(row=i, column=4).value = client_code[i-1]
    im_raw_data_tab.cell(row=i, column=4).value = im_raw_data_tab.cell(row=i, column=4).value.strip()
    im_raw_data_tab.cell(row=i, column=3).value = im_date[i-1]
    im_raw_data_tab.cell(row=i, column=3).value = im_raw_data_tab.cell(row=i, column=3).value.strip()
    im_raw_data_tab.cell(row=i, column=2).value = group_code[i - 1]
    im_raw_data_tab.cell(row=i, column=2).value = im_raw_data_tab.cell(row=i, column=2).value.strip()

# print (initial_margin_amount)
# print(group_code)

# print(open_file)

im_automationfile.save('\\\\altfps\\\\arcadiagroup$\Midoffice\IM Attachments\IM_Automationfile.xlsm')

conn = pyodbc.connect('DRIVER={SQL Server Native Client 11.0};SERVER=ArcSQL;DATABASE=RiskSandbox;TRUSTED_CONNECTION=yes')
cursor = conn.cursor()

im_automationfile = load_workbook('\\\\altfps\\arcadiagroup$\Midoffice\IM Attachments\IM_Automationfile.xlsm', keep_vba=True)
im_raw_data_tab = im_automationfile['IM_raw_data']
asofday = int(im_raw_data_tab.cell(row=2,column=3).value[:2])
asofmonth = int(im_raw_data_tab.cell(row=2,column=3).value[3:5])
asofyear = int(str('20' + im_raw_data_tab.cell(row=2,column=3).value[6:8]))
asofdate_time = datetime(asofyear, asofmonth, asofday)
no_rows = len(initial_margin_amount)

for i in range(2,no_rows):
    for j in range(2,4):
        asofdate_time
        if j==3:
            asofday = int(im_raw_data_tab.cell(row=i,column=j).value[:2])
            asofmonth = int(im_raw_data_tab.cell(row=i,column=j).value[3:5])
            asofyear = int(str('20' + im_raw_data_tab.cell(row=i,column=j).value[6:8]))
            asofdate_time = datetime(asofyear, asofmonth, asofday)
        elif j in [2,4,5]:
            group_code = im_raw_data_tab.cell(row=i, column=2).value
            clearing_account = im_raw_data_tab.cell(row=i, column=4).value
            value = im_raw_data_tab.cell(row=i, column=5).value
            if type(im_raw_data_tab.cell(i, 5).value) is not str:
                value = 0

    data = [asofdate_time, clearing_account, value, group_code, 'USD']
    print(data)


    cursor.execute("INSERT INTO InitialMarginInterest(AsOf, ClearingAccount, Value, GroupCode, Currency) VALUES(?,?,?,?,?)", data)
    cursor.commit()



cursor.close()
conn.close()
print('done')